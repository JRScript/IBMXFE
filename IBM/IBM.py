import requests
import re
import time
from openpyxl import Workbook
from openpyxl.styles import Font
import xlrd
import json
import base64
import os


if __name__ == "__main__":

	#Fetching API key Details
	cwd = os.getcwd()
	apikey = (cwd + '//API//api.xlsx')
	wb = xlrd.open_workbook(apikey)
	sheet = wb.sheet_by_index(0)
	sheet.cell_value(0, 0)

	key = sheet.cell_value(0, 1)
	password = sheet.cell_value(1, 1)

	#Encoding API Credentials
	cred = (key + ":" + password)
	encodedBytes = base64.b64encode(cred.encode("utf-8"))
	token = str(encodedBytes, "utf-8")
	headers = {'Authorization': "Basic " + token, 'Accept': 'application/json'}
	url = "https://api.xforce.ibmcloud.com"

	# Creating New Workbook for Results
	out_file = Workbook()
	ws_active =  out_file.active
	ws_active.title = ("CVE Details")

def IBM ():	

	#Adding Column Details 
	row = 1
	column = 1
	content = ["CVE-ID", "Type", "CVSSv3 Score", "CVSSv3 Vector", "Title", "Description", "Remedy", "Reported Time", "Platform Affected", "Exploitability", "Consequences", "Report Confidence", "Link Name", "Vendor Description", "Reference Links"]
	for item in content:
		ws_active.cell(row=row, column=column, value=item)
		ws_active.cell(row = row, column = column).font = Font(bold = True) 
		column += 1
	# Reding CVE from Input File and dumping JSON data from IBM
	input = (cwd + "//Input//CVE.xlsx")
	wb2 = xlrd.open_workbook(input)
	sheet1 = wb2.sheet_by_index(0)
	sheet1.cell_value(0, 0)
	r = 2
	c = 1
	j = 1
	for i in range (sheet1.nrows):
		cve_details = sheet1.cell_value(i, 0)
		apiurl = (url + "/vulnerabilities/search/" + cve_details)
		response = requests.get(apiurl, params='', headers=headers, timeout = 30)
		all_json = response.json()
		res = open(cwd + '//JSON//result.txt', 'w+')
		json.dump(all_json, res)
		res.close()
		data = open(cwd + '//JSON//result.txt', "r")
		contents = data.read()
		#CVE
		cve = re.findall("(?i)\"stdcode\"\:.*?\[\"(?P<cve>.*?)\"\]", contents)
		cve1 = str(','.join(cve))
		ws_active.cell(row=r, column=c, value=cve1)
		c += 1
		#Type
		Type = re.findall("(?i)\"type\"\:.*?\"(?P<Type>.*?)\"\,", contents)
		Type1 = str(','.join(Type))
		ws_active.cell(row=r, column=c, value=Type1)
		c += 1
		#CVSS
		cvss = re.findall("(?i)\"risk_level\"\:.*?(?P<CVSS>.*?)\,.*?", contents)
		cvss1 = str(','.join(cvss))
		ws_active.cell(row=r, column=c, value=cvss1)
		c += 1
		#CVSSv3 Calculator and Vector
			# AccessVector
		AV = re.findall("(?i)\"access_vector\"\:.*?\"(?P<AV>.*?)\"\,.*?", contents)
		AV1 = str(','.join(AV))
		AV2 = AV1[:1]
			# Access Complexity
		AC = re.findall("(?i)\"access_complexity\"\:.*?\"(?P<AC>.*?)\"\,.*?", contents)
		AC1 = str(','.join(AC))
		AC2 = AC1[:1]
			# Privilege Required
		PR = re.findall("(?i)\"privilegesrequired\"\:.*?\"(?P<CVSS>.*?)\"\,.*?", contents)
		PR1 = str(','.join(PR))
		PR2 = PR1[:1]
			# User Interaction
		UI = re.findall("(?i)\"userinteraction\"\:.*?\"(?P<UI>.*?)\"\,.*?", contents)
		UI1 = str(','.join(UI))
		UI2 = UI1[:1]
			# Scope
		SC = re.findall("(?i)\"scope\"\:.*?\"(?P<SC>.*?)\"\,.*?", contents)
		SC1 = str(','.join(SC))
		SC2 = SC1[:1]
			#Confidentiality
		CO = re.findall("(?i)\"confidentiality_impact\"\:.*?\"(?P<CVSS>.*?)\"\,.*?", contents)
		CO1 = str(','.join(CO))
		CO2 = CO1[:1]
			#Integrity
		IN = re.findall("(?i)\"integrity_impact\"\:.*?\"(?P<CVSS>.*?)\"\,.*?", contents)
		IN1 = str(','.join(IN))
		IN2 = IN1[:1]
			#Availability
		AL = re.findall("(?i)\"availability_impact\"\:.*?\"(?P<CVSS>.*?)\"\,.*?", contents)
		AL1 = str(','.join(AL))
		AL2 = AL1[:1]
			#CVSSv3 Vector
		cvssc = ("CVSS:3.0/AV:" + AV2 + "/AC:" + AC2 + "/PR:" + PR2 + "/UI:" + UI2 + "/S:" + SC2 + "/C:" + CO2 + "/I:" + IN2 + "/A:" + AL2)
		ws_active.cell(row=r, column=c, value=cvssc)
		c += 1
		#Title
		title = re.findall("(?i)\"title\"\:.*?\"(?P<Title>.*?)\"\,", contents)
		title1 = (str(','.join(title)))
		ws_active.cell(row=r, column=c, value=title1)
		c += 1
		#Description
		des = re.findall("(?i)title.*?\"description\"\:.*?\"(?P<Des>.*?)\"\,*?", contents)
		des1 = (str(','.join(des)))
		ws_active.cell(row=r, column=c, value=des1)
		c += 1
		#Remedy
		rem = re.findall("(?i)\"Remedy\"\:.*?\"(?P<Rem>.*?)\"\,*?", contents)
		rem1 = (str(','.join(rem)))
		ws_active.cell(row=r, column=c, value=rem1)
		c += 1
		#ReportedTime
		rt = re.findall("(?i)\"Reported\"\:.*?\"(?P<rt>.*?)\"\,*?", contents)
		rt1 = (str(','.join(rt)))
		ws_active.cell(row=r, column=c, value=rt1)
		c += 1
		#PlatformsAffected
		pa = re.findall("(?i)\"platforms_affected\"\:.*?\[(?P<pa>.*?)\],.*?\"exploitability.*?", contents)
		pa1 = (str(','.join(pa)))
		ws_active.cell(row=r, column=c, value=pa1)
		c += 1
		#Exploitability
		ex = re.findall("(?i)\"exploitability\"\:.*?\"(?P<Socre>.*?)\".*?", contents)
		ex1 = (str(','.join(ex)))
		ws_active.cell(row=r, column=c, value=ex1)
		c += 1
		#Consequences
		con = re.findall("(?i)\"consequences\"\:.*?\"(?P<Socre>.*?)\".*?", contents)
		con1 = (str(','.join(con)))
		ws_active.cell(row=r, column=c, value=con1)
		c += 1
		#ReportConfidence
		rc = re.findall("(?i)\"report\_confidence\"\:.*?\"(?P<Socre>.*?)\".*?", contents)
		rc1 = (str(','.join(rc)))
		ws_active.cell(row=r, column=c, value=rc1)
		c += 1
		#LinkName
		ln = re.findall("(?i)\"link_name\"\:.*?\"(?P<Socre>.*?)\".*?", contents)
		ln1 = (str(ln)[1:-1])
		ws_active.cell(row=r, column=c, value=ln1)
		c += 1
		#VendorDescription
		vd = re.findall("(?i)\"exploitability\".*\"description.*?\:.*?\"(?P<des1>.*?)\".*?description.*?\:.*?\"(?P<des2>.*?)\".*?", contents)
		vd1 = ((str(vd)[2:-2]))
		ws_active.cell(row=r, column=c, value=vd1)
		c += 1
		#TaregtLinks
		tl = re.findall("(?i)\"exploitability\".*\"link_target.*?\:.*?\"(?P<Target1>.*?)\".*?\"link_target.*?\:.*?\"(?P<Target2>.*?)\".*?", contents)
		tl1  = ((str(tl)[2:-2]))
		ws_active.cell(row=r, column=c, value=tl1)

		x = (sheet1.nrows)
		print (j, "/" ,x)
		j += 1
		r += 1
		c = 1
IBM()

out_file.save(filename = cwd + "//Results//CVE-" + time.strftime("%m-%d-%Y") + ".xlsx")
print ("Successfully Generated the CVE Report")